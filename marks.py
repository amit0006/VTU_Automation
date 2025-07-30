import sys
import os
import re
import pandas as pd
from doctr.io import DocumentFile
from doctr.models import ocr_predictor
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import certifi
os.environ['SSL_CERT_FILE'] = certifi.where()

# ========== Step 0: Handle CLI Argument ==========
if len(sys.argv) != 2:
    print("❌ Usage: python marks.py <screenshot_path>")
    sys.exit(1)

image_path = sys.argv[1]
if not os.path.exists(image_path):
    print(f"❌ File not found: {image_path}")
    sys.exit(1)

filename = os.path.basename(image_path)

# ========== Step 1: Load OCR Model ==========
model = ocr_predictor(pretrained=True)

# ========== Step 2: Setup Excel Workbook ==========
excel_filename = "vtu_structured_results.xlsx"
if os.path.exists(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["University Seat Number"])  # Initial header

# ========== Step 3: Build Header Index ==========
header_cells = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
existing_usns = {ws.cell(row=i, column=header_cells["University Seat Number"]).value for i in range(2, ws.max_row + 1)}

# ========== Step 4: OCR Processing ==========
doc = DocumentFile.from_images(image_path)
result = model(doc)
json_output = result.export()

# Extract all text lines
all_lines = []
for page in json_output['pages']:
    for block in page['blocks']:
        for line in block['lines']:
            line_text = ' '.join(word['value'] for word in line['words'])
            all_lines.append(line_text.strip())

full_text = '\n'.join(all_lines)

# Extract USN
usn_match = re.search(r'University\s*Seat\s*Number\s*[:\-]?\s*([0-9A-Z]{8,12})', full_text, re.IGNORECASE)
usn = usn_match.group(1).strip() if usn_match else "Not Found"

if usn in existing_usns or usn == "Not Found":
    print(f"⏭️ Skipping {filename} (Already processed or USN not found)")
    sys.exit(0)

# ========== Step 5: Extract subject marks and results ==========
subjects = []
i = 0
while i < len(all_lines):
    line = all_lines[i].strip()
    # Updated regex to be more flexible with subject code format
    code_match = re.match(r'^([A-Z]{2,}\d{2,}[A-Z]?\d*|[0-9]{2}[A-Z]{2,}\d{1,})', line)

    if code_match:
        code = code_match.group(1)
        internal = external = total = None
        result_status = ""

        # Look for result status in the current line or next few lines
        possible_status_lines = all_lines[i:i+5] # Check current line and next 4 lines
        for l in possible_status_lines:
            status_match = re.search(r'\b(P|F|A|W)\b', l.strip()) # Look for P, F, A, W as whole words
            if status_match:
                result_status = status_match.group(1)
                break

        # Attempt to find marks in the lines following the subject code
        int_line = all_lines[i + 1] if i + 1 < len(all_lines) else ''
        ext_line = all_lines[i + 2] if i + 2 < len(all_lines) else ''

        int_vals = re.findall(r'\d+', int_line)
        if int_vals:
            internal = int(int_vals[0])

        ext_vals = re.findall(r'\d+', ext_line)
        if len(ext_vals) >= 2:
            external = int(ext_vals[0])
            total = int(ext_vals[1])
        elif len(ext_vals) == 1:
            external = int(ext_vals[0])
            total = internal + external if internal is not None else external
        # If no external marks are found, total might still be internal
        else:
            external = None # Set to None explicitly if not found
            total = internal if internal is not None else None


        # Refined logic for total if it exceeds expected range or needs recalculation
        if total is not None and total > 200 and internal is not None and external is not None:
            total = internal + external
        elif total is None and internal is not None and external is not None:
            total = internal + external
        elif total is None and internal is not None:
             total = internal # If only internal is found, assume total is internal if external is missing

        # --- MODIFICATION START ---
        # Add subject if a code is found, regardless of whether marks are extracted.
        # This ensures 'A' subjects are recorded even without numerical marks.
        subjects.append({
            "Subject Code": code,
            "Total": total, # Will be None if marks aren't found, which is fine for 'A'
            "Result": result_status
        })
        i += 3 # Move past current subject code, internal, and external lines (even if empty)
        continue # Continue to next iteration
        # --- MODIFICATION END ---
    i += 1

# Deduplicate subjects and prioritize entries with a 'Total' mark
subjects_cleaned = {}
for sub in subjects:
    code = sub["Subject Code"]
    # Prioritize if it has a result status or higher total
    # Also prioritize if it has a total, or a result status (like 'A') if no total is present
    if code not in subjects_cleaned or \
       (sub["Result"] and not subjects_cleaned[code]["Result"]) or \
       (sub["Total"] is not None and (subjects_cleaned[code]["Total"] is None or sub["Total"] > subjects_cleaned[code]["Total"])) or \
       (sub["Result"] == 'A' and subjects_cleaned[code]["Total"] is None): # Added condition for 'A' with no total
        subjects_cleaned[code] = sub


# Add new subject columns if any
for code in list(subjects_cleaned.keys()):
    if code not in header_cells:
        new_col_total = ws.max_column + 1
        ws.cell(row=1, column=new_col_total).value = code
        header_cells[code] = new_col_total


# Build row
row = [None] * (ws.max_column) # Ensure row size matches current max columns
row[header_cells["University Seat Number"] - 1] = usn

for code, sub_info in subjects_cleaned.items():
    if code in header_cells:
        # If total is None, leave cell blank or put a placeholder like "N/A"
        # For coloring, we just need the cell to exist.
        row[header_cells[code] - 1] = sub_info["Total"] # This will be None if no marks found


# Append row
ws.append(row)
last_row = ws.max_row

# Define fill colors
fill_red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid") # Red
fill_green = PatternFill(start_color="FF00FF00", end_color="FF00FF00", fill_type="solid") # Green

# Apply colors based on result status
for code, sub_info in subjects_cleaned.items():
    if code in header_cells: # Ensure the subject column exists
        col_idx = header_cells[code]
        cell_to_color = ws.cell(row=last_row, column=col_idx)

        # Reset fill for the cell first to ensure it's white/no-fill by default
        cell_to_color.fill = PatternFill(fill_type=None) # Clears any previous fill

        if sub_info["Result"] == "F":
            cell_to_color.fill = fill_red
        elif sub_info["Result"] == "A":
            cell_to_color.fill = fill_green
        # For 'P' and 'W', the fill remains cleared (white).

print(f"✅ Processed: {filename} - {usn}")

# Save workbook
wb.save(excel_filename)
print("📘 Excel file saved:", excel_filename)